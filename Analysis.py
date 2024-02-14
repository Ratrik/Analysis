from openpyxl import Workbook
import re
from tkinter import *
from tkinter import ttk


# Интерфейс


def finish():
    root.destroy()  # ручное закрытие окна и всего приложения


# noinspection PyTypeChecker
def click():
    """ Формирование документа, импорт данных и статистические расчеты """

    name: str = entry_1.get()
    period: str = entry_2.get()

    # Импорт

    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Год'
    ws['B1'] = 'Город'
    ws['C1'] = 'Зоопарк'
    ws['D1'] = 'Численность'
    ws['H1'] = 'Родившиеся'
    ws['L1'] = 'Павшие'
    ws['D2'] = 'Самцы'
    ws['E2'] = 'Самки'
    ws['F2'] = 'Неизвестно'
    ws['G2'] = 'Общее'
    ws['H2'] = 'Самцы'
    ws['I2'] = 'Самки'
    ws['J2'] = 'Неизвестно'
    ws['K2'] = 'Общее'
    ws['L2'] = 'Самцы'
    ws['M2'] = 'Самки'
    ws['N2'] = 'Неизвестно'
    ws['O2'] = 'Общее'

    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:G1')
    ws.merge_cells('H1:K1')
    ws.merge_cells('L1:O1')

    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['J'].width = 13
    ws.column_dimensions['N'].width = 13

    name = name.replace(' ', '')

    period = re.split('-', period)

    years = []

    for i in range(int(period[0]), int(period[1]) + 1):
        years.append(i)

    # noinspection PyTypeChecker
    def form(x, y):

        for i in range(1, 1000):
            if i == 4 and ws['B' + str(i)].value is None:
                index = 4
                ind = 3
            if i > 3 and ws['B' + str(i)].value is not None and ws['B' + str(i + 1)].value is None:
                index = i + 5
                ind = i + 4

        # Численность
        sl_index = index
        ws['A' + str(ind)] = x
        d = open('N' + str(x) + '.txt', 'r', encoding='utf-8')
        g = str(d.read())
        g = re.split('\n', g)
        digits = '1234567890'
        low = 'q w e r t y u i o p a s d f g h j k l z x c v b m'
        low = re.split(' ', low)

        for i in range(1, len(g)):

            for j in range(len(low)):
                if low[j] in g[i]:
                    g[i] = '!' + g[i] + '@'
                    break

        g = '\n'.join(g)
        g = g.replace('\n', '')
        g = re.split('!', g)

        for i in range(len(g)):
            if y in g[i]:
                indit = 1
                spec = g[i]
                spec = re.split('@', spec)
                spec[1] = re.split(';', spec[1])
                lis = '1234567890'

                for i in range(len(spec[1])):
                    ed = []

                    for j in range(len(spec[1][i])):
                        ed.append(spec[1][i][j])

                    for k in range(len(ed)):
                        if ed[k] in lis:
                            ed[k] = '/' + ed[k]
                            break

                    spec[1][i] = ''.join(ed)
                    spec[1][i] = re.split('/', spec[1][i])
                towns = spec[1]
                tind = 0

                for i in range(len(towns)):
                    if '(' in towns[i][0]:
                        towns[i][0] = re.split('(', towns[i][0])
                        towns[i][0][1] = towns[i][0][1].replace(')', '')
                        index = str(index)
                        ws['B' + index] = towns[i][0][0]
                        ws['C' + index] = towns[i][0][1]
                        index = int(index)
                        index += 1
                        tind += 1
                    elif '«' in towns[i][0]:
                        towns[i][0] = re.split('«', towns[i][0])
                        towns[i][0][1] = towns[i][0][1].replace('»', '')
                        index = str(index)
                        ws['B' + index] = towns[i][0][0]
                        ws['C' + index] = towns[i][0][1]
                        index = int(index)
                        index += 1
                        tind += 1
                    else:
                        index = str(index)
                        ws['B' + index] = towns[i][0]
                        index = int(index)
                        index += 1
                        tind += 1
                index = index - tind

                for i in range(len(towns)):
                    if len(towns[i]) == 2:
                        index = str(index)
                        ws['F' + index] = int(towns[i][1])
                        index = int(index)
                        index += 1
                    elif len(towns[i]) == 3:
                        index = str(index)
                        ws['D' + index] = int(towns[i][1])
                        ws['E' + index] = int(towns[i][2])
                        index = int(index)
                        index += 1
                    elif len(towns[i]) == 4:
                        index = str(index)
                        ws['D' + index] = int(towns[i][1])
                        ws['E' + index] = int(towns[i][2])
                        ws['F' + index] = int(towns[i][3])
                        index = int(index)
                        index += 1
        # Рождаемость

        d = open('B' + str(x) + '.txt', 'r', encoding='utf-8')
        g = str(d.read())
        g = re.split('\n', g)
        digits = '1234567890'
        low = 'q w e r t y u i o p a s d f g h j k l z x c v b m'
        low = re.split(' ', low)

        for i in range(1, len(g)):
            for j in range(len(low)):
                if low[j] in g[i]:
                    g[i] = '!' + g[i] + '@'
                    break

        g = '\n'.join(g)
        g = g.replace('\n', '')
        g = re.split('!', g)

        for i in range(len(g)):
            if y in g[i]:
                spec = g[i]
                spec = re.split('@', spec)
                spec[1] = re.split(';', spec[1])
                lis = '1234567890'

                for i in range(len(spec[1])):
                    ed = []

                    for j in range(len(spec[1][i])):
                        ed.append(spec[1][i][j])

                    for k in range(len(ed)):
                        if ed[k] in lis:
                            ed[k] = '/' + ed[k]
                            break
                        elif ed[k] == '(' and ed[k + 1] in lis:
                            ed[k] = '/' + ed[k]
                            break

                    for l in range(len(ed)):
                        if ed[l] == '(' and ed[l + 1] in lis:
                            ed[l] = '/' + ed[l]
                            break
                    spec[1][i] = ''.join(ed)
                    spec[1][i] = re.split('/', spec[1][i])
                    # print(spec[1][i])
                towns = spec[1]

                for i in range(sl_index, index):

                    for j in range(len(towns)):
                        istr = str(i)
                        if ws.cell(row=i, column=2).value in towns[j][0]:
                            if len(towns[j]) == 2:
                                if '(' in towns[j][1]:
                                    towns[j][1] = towns[j][1].replace('(', '')
                                    towns[j][1] = towns[j][1].replace(')', '')
                                    ws['N' + istr] = int(towns[j][1])
                                else:
                                    ws['J' + istr] = int(towns[j][1])
                            elif len(towns[j]) == 3:
                                if '(' in towns[j][1]:
                                    towns[j][1] = towns[j][1].replace('(', '')
                                    towns[j][2] = towns[j][2].replace(')', '')
                                    ws['L' + istr] = int(towns[j][1])
                                    ws['M' + istr] = int(towns[j][2])
                                elif '(' in towns[j][2]:
                                    towns[j][2] = towns[j][2].replace('(', '')
                                    towns[j][2] = towns[j][2].replace(')', '')
                                    ws['J' + istr] = int(towns[j][1])
                                    ws['N' + istr] = int(towns[j][2])
                                else:
                                    ws['H' + istr] = int(towns[j][1])
                                    ws['I' + istr] = int(towns[j][2])
                            elif len(towns[j]) == 4:
                                if '(' in towns[j][3]:
                                    towns[j][3] = towns[j][3].replace('(', '')
                                    towns[j][3] = towns[j][3].replace(')', '')
                                    ws['H' + istr] = int(towns[j][1])
                                    ws['I' + istr] = int(towns[j][2])
                                    ws['N' + istr] = int(towns[j][3])
                                elif '(' in towns[j][2]:
                                    towns[j][2] = towns[j][2].replace('(', '')
                                    towns[j][3] = towns[j][3].replace(')', '')
                                    ws['J' + istr] = int(towns[j][1])
                                    ws['K' + istr] = int(towns[j][2])
                                    ws['L' + istr] = int(towns[j][3])
                                elif '(' in towns[j][1]:
                                    towns[j][1] = towns[j][1].replace('(', '')
                                    towns[j][3] = towns[j][3].replace(')', '')
                                    ws['L' + istr] = int(towns[j][1])
                                    ws['M' + istr] = int(towns[j][2])
                                    ws['N' + istr] = int(towns[j][3])
                                else:
                                    ws['H' + istr] = int(towns[j][1])
                                    ws['I' + istr] = int(towns[j][2])
                                    ws['J' + istr] = int(towns[j][3])
                            elif len(towns[j]) == 5:
                                if '(' in towns[j][4]:
                                    towns[j][4] = towns[j][4].replace('(', '')
                                    towns[j][4] = towns[j][4].replace(')', '')
                                    ws['H' + istr] = int(towns[j][1])
                                    ws['I' + istr] = int(towns[j][2])
                                    ws['J' + istr] = int(towns[j][3])
                                    ws['N' + istr] = int(towns[j][4])
                                elif '(' in towns[j][3]:
                                    towns[j][3] = towns[j][3].replace('(', '')
                                    towns[j][4] = towns[j][4].replace(')', '')
                                    ws['H' + istr] = int(towns[j][1])
                                    ws['I' + istr] = int(towns[j][2])
                                    ws['L' + istr] = int(towns[j][3])
                                    ws['M' + istr] = int(towns[j][4])
                                elif '(' in towns[j][2]:
                                    towns[j][2] = towns[j][2].replace('(', '')
                                    towns[j][4] = towns[j][4].replace(')', '')
                                    ws['J' + istr] = int(towns[j][1])
                                    ws['L' + istr] = int(towns[j][2])
                                    ws['M' + istr] = int(towns[j][3])
                                    ws['N' + istr] = int(towns[j][4])
                            elif len(towns[j]) == 6:
                                if '(' in towns[j][4]:
                                    towns[j][4] = towns[j][4].replace('(', '')
                                    towns[j][5] = towns[j][5].replace(')', '')
                                    ws['H' + istr] = int(towns[j][1])
                                    ws['I' + istr] = int(towns[j][2])
                                    ws['J' + istr] = int(towns[j][3])
                                    ws['L' + istr] = int(towns[j][4])
                                    ws['M' + istr] = int(towns[j][5])
                                elif '(' in towns[j][3]:
                                    towns[j][3] = towns[j][3].replace('(', '')
                                    towns[j][5] = towns[j][5].replace(')', '')
                                    ws['H' + istr] = int(towns[j][1])
                                    ws['I' + istr] = int(towns[j][2])
                                    ws['L' + istr] = int(towns[j][3])
                                    ws['M' + istr] = int(towns[j][4])
                                    ws['N' + istr] = int(towns[j][5])
                            elif len(towns[j]) == 7:
                                towns[j][4] = towns[j][4].replace('(', '')
                                towns[j][6] = towns[j][6].replace(')', '')
                                ws['H' + istr] = int(towns[j][1])
                                ws['I' + istr] = int(towns[j][2])
                                ws['J' + istr] = int(towns[j][3])
                                ws['L' + istr] = int(towns[j][4])
                                ws['M' + istr] = int(towns[j][5])
                                ws['N' + istr] = int(towns[j][6])

                sl_index = index
                print('Конец')

    for i in range(len(years)):
        form(years[i], name)

    # Заполнение формы

    ws['R1'] = 'Статистика'
    ws.merge_cells('R1:DB1')

    # Количество особей

    ws['R2'] = 'Количество особей'
    ws['R3'] = 'Год'
    ws['S3'] = 'Численность'
    ws['W3'] = 'Родившиеся'
    ws['AA3'] = 'Павшие'
    ws['AB3'] = 'Город'
    ws['AC3'] = 'Частота'
    ws['S4'] = 'Самцы'
    ws['T4'] = 'Самки'
    ws['U4'] = 'Неизвестно'
    ws['V4'] = 'Общее'
    ws['W4'] = 'Самцы'
    ws['X4'] = 'Самки'
    ws['Y4'] = 'Неизвестно'
    ws['Z4'] = 'Общее'
    ws.merge_cells('R2:AC2')
    ws.merge_cells('R3:R4')
    ws.merge_cells('S3:V3')
    ws.merge_cells('W3:Z3')
    ws.merge_cells('AA3:AA4')
    ws.merge_cells('AB3:AB4')
    ws.merge_cells('AC3:AC4')
    ws.column_dimensions['AB'].width = 20
    ws.column_dimensions['U'].width = 13
    ws.column_dimensions['Y'].width = 13

    # Средняя арифметическая

    ws['AE2'] = 'Средняя арифметическая'
    ws['AE3'] = 'Год'
    ws['AF3'] = 'Численность'
    ws['AJ3'] = 'Родившиеся'
    ws['AN3'] = 'Павшие'
    ws['AF4'] = 'Самцы'
    ws['AG4'] = 'Самки'
    ws['AH4'] = 'Неизвестно'
    ws['AI4'] = 'Общее'
    ws['AJ4'] = 'Самцы'
    ws['AK4'] = 'Самки'
    ws['AL4'] = 'Неизвестно'
    ws['AM4'] = 'Общее'
    ws.merge_cells('AE2:AN2')
    ws.merge_cells('AE3:AE4')
    ws.merge_cells('AF3:AI3')
    ws.merge_cells('AJ3:AM3')
    ws.merge_cells('AN3:AN4')
    ws.column_dimensions['AH'].width = 13
    ws.column_dimensions['AL'].width = 13

    # Средняя геометрическая

    ws['AP2'] = 'Средняя геометрическая'
    ws['AP3'] = 'Год'
    ws['AQ3'] = 'Численность'
    ws['AU3'] = 'Родившиеся'
    ws['AY3'] = 'Павшие'
    ws['AQ4'] = 'Самцы'
    ws['AR4'] = 'Самки'
    ws['AS4'] = 'Неизвестно'
    ws['AT4'] = 'Общее'
    ws['AU4'] = 'Самцы'
    ws['AV4'] = 'Самки'
    ws['AW4'] = 'Неизвестно'
    ws['AX4'] = 'Общее'
    ws.merge_cells('AP2:AY2')
    ws.merge_cells('AP3:AP4')
    ws.merge_cells('AQ3:AT3')
    ws.merge_cells('AU3:AX3')
    ws.merge_cells('AY3:AY4')
    ws.column_dimensions['AP'].width = 13
    ws.column_dimensions['AS'].width = 13
    ws.column_dimensions['AW'].width = 13

    # Мода

    ws['BA2'] = 'Мода'
    ws['BA3'] = 'Год'
    ws['BB3'] = 'Численность'
    ws['BF3'] = 'Родившиеся'
    ws['BJ3'] = 'Павшие'
    ws['BB4'] = 'Самцы'
    ws['BC4'] = 'Самки'
    ws['BD4'] = 'Неизвестно'
    ws['BE4'] = 'Общее'
    ws['BF4'] = 'Самцы'
    ws['BG4'] = 'Самки'
    ws['BH4'] = 'Неизвестно'
    ws['BI4'] = 'Общее'
    ws.merge_cells('BA2:BJ2')
    ws.merge_cells('BA3:BA4')
    ws.merge_cells('BB3:BE3')
    ws.merge_cells('BF3:BI3')
    ws.merge_cells('BJ3:BJ4')
    ws.column_dimensions['BD'].width = 13
    ws.column_dimensions['BH'].width = 13

    # Дисперсия

    ws['BL2'] = 'Дисперсия'
    ws['BL3'] = 'Год'
    ws['BM3'] = 'Численность'
    ws['BQ3'] = 'Родившиеся'
    ws['BU3'] = 'Павшие'
    ws['BM4'] = 'Самцы'
    ws['BN4'] = 'Самки'
    ws['BO4'] = 'Неизвестно'
    ws['BP4'] = 'Общее'
    ws['BQ4'] = 'Самцы'
    ws['BR4'] = 'Самки'
    ws['BS4'] = 'Неизвестно'
    ws['BT4'] = 'Общее'
    ws.merge_cells('BL2:BU2')
    ws.merge_cells('BL3:BL4')
    ws.merge_cells('BM3:BP3')
    ws.merge_cells('BQ3:BT3')
    ws.merge_cells('BU3:BU4')
    ws.column_dimensions['BO'].width = 13
    ws.column_dimensions['BS'].width = 13

    # Среднее квадратичное отклонение

    ws['BW2'] = 'Среднее квадратичное отклонение'
    ws['BW3'] = 'Год'
    ws['BX3'] = 'Численность'
    ws['CB3'] = 'Родившиеся'
    ws['CF3'] = 'Павшие'
    ws['BX4'] = 'Самцы'
    ws['BY4'] = 'Самки'
    ws['BZ4'] = 'Неизвестно'
    ws['CA4'] = 'Общее'
    ws['CB4'] = 'Самцы'
    ws['CC4'] = 'Самки'
    ws['CD4'] = 'Неизвестно'
    ws['CE4'] = 'Общее'
    ws.merge_cells('BW2:CF2')
    ws.merge_cells('BW3:BW4')
    ws.merge_cells('BX3:CA3')
    ws.merge_cells('CB3:CE3')
    ws.merge_cells('CF3:CF4')
    ws.column_dimensions['BZ'].width = 13
    ws.column_dimensions['CD'].width = 13

    # Коэффициент вариации

    ws['CH2'] = 'Коэффициент вариации, %'
    ws['CH3'] = 'Год'
    ws['CI3'] = 'Численность'
    ws['CM3'] = 'Родившиеся'
    ws['CQ3'] = 'Павшие'
    ws['CI4'] = 'Самцы'
    ws['CJ4'] = 'Самки'
    ws['CK4'] = 'Неизвестно'
    ws['CL4'] = 'Общее'
    ws['CM4'] = 'Самцы'
    ws['CN4'] = 'Самки'
    ws['CO4'] = 'Неизвестно'
    ws['CP4'] = 'Общее'
    ws.merge_cells('CH2:CQ2')
    ws.merge_cells('CH3:CH4')
    ws.merge_cells('CI3:CL3')
    ws.merge_cells('CM3:CP3')
    ws.merge_cells('CQ3:CQ4')
    ws.column_dimensions['CK'].width = 13
    ws.column_dimensions['CO'].width = 13

    # Общее распределение

    ws['CS2'] = 'Общее распределение'
    ws['CS3'] = 'Отклонение'
    ws['CT3'] = 'Численность'
    ws['CX3'] = 'Родившиеся'
    ws['DB3'] = 'Павшие'
    ws['CT4'] = 'Самцы'
    ws['CU4'] = 'Самки'
    ws['CV4'] = 'Неизвестно'
    ws['CW4'] = 'Общее'
    ws['CX4'] = 'Самцы'
    ws['CY4'] = 'Самки'
    ws['CZ4'] = 'Неизвестно'
    ws['DA4'] = 'Общее'
    ws.merge_cells('CS2:DB2')
    ws.merge_cells('CS3:CS4')
    ws.merge_cells('CT3:CW3')
    ws.merge_cells('CX3:DA3')
    ws.merge_cells('DB3:DB4')
    ws.column_dimensions['CS'].width = 13
    ws.column_dimensions['CV'].width = 13
    ws.column_dimensions['CZ'].width = 13

    # Расчет

    file_name = name
    years = period
    cell_list = []
    years_list = []
    years[0] = int(years[0])
    years[1] = int(years[1])

    for i in range(years[0], years[1] + 1):
        years_list.append(i)
    years_list.append('Всего')
    print(years_list)

    for i in range(3, 700):
        index = str(i)
        if ws['A' + index].value is not None:
            cell_list.append(i)

    cell_list.append(700)
    print(cell_list)

    # Расчет суммы

    for i in range(len(cell_list)):
        if cell_list[i] != 700:

            # Численность
            for j in range(cell_list[i] + 1, cell_list[i + 1] - 3):
                number = str(j)
                if ws['D' + number].value is not None:
                    ws['G' + number] = ws['D' + number].value

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['E' + number].value is not None:
                    if ws['G' + number].value is not None:
                        ws['G' + number] = ws['G' + number].value + ws['E' + number].value
                    if ws['G' + number].value is None:
                        ws['G' + number] = ws['E' + number].value

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['F' + number].value is not None:
                    if ws['G' + number].value is not None:
                        ws['G' + number] = ws['G' + number].value + ws['F' + number].value
                    if ws['G' + number].value is None:
                        ws['G' + number] = ws['F' + number].value

            # Рождаемость
            for j in range(cell_list[i] + 1, cell_list[i + 1] - 3):
                number = str(j)
                if ws['H' + number].value is not None:
                    ws['K' + number] = ws['H' + number].value

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['I' + number].value is not None:
                    if ws['K' + number].value is not None:
                        ws['K' + number] = ws['K' + number].value + ws['I' + number].value
                    if ws['K' + number].value is None:
                        ws['K' + number] = ws['I' + number].value

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['J' + number].value is not None:
                    if ws['K' + number].value is not None:
                        ws['K' + number] = ws['K' + number].value + ws['J' + number].value
                    if ws['K' + number].value is None:
                        ws['K' + number] = ws['J' + number].value
            # Павшие
            for j in range(cell_list[i] + 1, cell_list[i + 1] - 3):
                number = str(j)
                if ws['L' + number].value is not None:
                    ws['O' + number] = ws['L' + number].value

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['M' + number].value is not None:
                    if ws['O' + number].value is not None:
                        ws['O' + number] = ws['O' + number].value + ws['M' + number].value
                    if ws['O' + number].value is None:
                        ws['O' + number] = ws['M' + number].value

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['N' + number].value is not None:
                    if ws['O' + number].value is not None:
                        ws['O' + number] = ws['O' + number].value + ws['N' + number].value
                    if ws['O' + number].value is None:
                        ws['O' + number] = ws['N' + number].value

    # Средняя арифметическая

    # Численность
    count = 5

    for i in range(len(years_list)):
        count_str = str(count)
        ws['AE' + count_str] = years_list[i]
        ws['R' + count_str] = years_list[i]
        count += 1
        if years_list[i] == 'Всего':
            ws['R' + count_str] = None

    males_list = []
    females_list = []
    unknown_list = []
    average_males_year = []
    average_females_year = []
    average_unknown_year = []

    for i in range(len(cell_list)):
        if cell_list[i] != 700:
            males = 0
            females = 0
            unknown = 0
            count_females = 0
            count_unknown = 0
            count_males = 0

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['D' + number].value is not None:
                    males += ws['D' + number].value
                    count_males += 1

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['E' + number].value is not None:
                    females += ws['E' + number].value
                    count_females += 1

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['F' + number].value is not None:
                    unknown += ws['F' + number].value
                    count_unknown += 1

        x = int
        y = int
        z = int

        if count_males != 0:
            x = males / count_males

        if count_males == 0:
            x = 0

        if count_females != 0:
            y = females / count_females

        if count_females == 0:
            y = 0

        if count_unknown != 0:
            z = unknown / count_unknown

        if count_unknown == 0:
            z = 0

        average_males_year.append(x)
        average_females_year.append(y)
        average_unknown_year.append(z)
        males_list.append(males)
        females_list.append(females)
        unknown_list.append(unknown)

    # Родившиеся
    born_males_list = []
    born_females_list = []
    born_unknown_list = []
    average_born_males_year = []
    average_born_females_year = []
    average_born_unknown_year = []

    for i in range(len(cell_list)):
        if cell_list[i] != 700:
            born_males = 0
            born_females = 0
            born_unknown = 0
            count_born_females = 0
            count_born_unknown = 0
            count_born_males = 0

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['H' + number].value is not None:
                    born_males += ws['H' + number].value
                    count_born_males += 1

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['I' + number].value is not None:
                    born_females += ws['I' + number].value
                    count_born_females += 1

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['J' + number].value is not None:
                    born_unknown += ws['J' + number].value
                    count_born_unknown += 1
        x = int
        y = int
        z = int

        if count_born_males != 0:
            x = born_males / count_born_males

        if count_born_males == 0:
            x = 0

        if count_born_females != 0:
            y = born_females / count_born_females

        if count_born_females == 0:
            y = 0

        if count_born_unknown != 0:
            z = born_unknown / count_born_unknown

        if count_born_unknown == 0:
            z = 0

        average_born_males_year.append(x)
        average_born_females_year.append(y)
        average_born_unknown_year.append(z)
        born_males_list.append(born_males)
        born_females_list.append(born_females)
        born_unknown_list.append(born_unknown)

    # Павшие
    fallen_list = []
    average_fallen_year = []

    for i in range(len(cell_list)):
        if cell_list[i] != 700:
            fallen = 0
            count_fallen = 0
            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['L' + number].value is not None:
                    fallen += ws['L' + number].value
                    count_fallen += 1
            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['M' + number].value is not None:
                    fallen += ws['M' + number].value
                    count_fallen += 1
            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['N' + number].value is not None:
                    fallen += ws['N' + number].value
                    count_fallen += 1
        x = int
        if count_fallen != 0:
            x = fallen / count_fallen
        if count_fallen == 0:
            x = 0
        average_fallen_year.append(x)
        fallen_list.append(fallen)

    # Заполнение
    for i in range(len(years_list) - 1):
        number = i + 5
        number = str(number)
        num = 0
        b_num = 0
        ws['S' + number] = males_list[i]
        ws['T' + number] = females_list[i]
        ws['U' + number] = unknown_list[i]
        ws['V' + number] = males_list[i] + females_list[i] + unknown_list[i]
        ws['W' + number] = born_males_list[i]
        ws['X' + number] = born_females_list[i]
        ws['Y' + number] = born_unknown_list[i]
        ws['Z' + number] = born_males_list[i] + born_females_list[i] + born_unknown_list[i]
        ws['AA' + number] = fallen_list[i]
        ws['AF' + number] = average_males_year[i]
        ws['AG' + number] = average_females_year[i]
        ws['AH' + number] = average_unknown_year[i]
        ws['AJ' + number] = average_born_males_year[i]
        ws['AK' + number] = average_born_females_year[i]
        ws['AL' + number] = average_born_unknown_year[i]
        ws['AN' + number] = average_fallen_year[i]

        # Общая средняя арифметическая
        if males_list[i] > 0:
            num += 1
        if females_list[i] > 0:
            num += 1
        if unknown_list[i] > 0:
            num += 1
        if num != 0:
            ws['AI' + number] = (average_males_year[i] + average_females_year[i] + average_unknown_year[i]) / num
        if num == 0:
            ws['AI' + number] = 0
        if born_males_list[i] > 0:
            b_num += 1
        if born_females_list[i] > 0:
            b_num += 1
        if born_unknown_list[i] > 0:
            b_num += 1
        if b_num != 0:
            ws['AM' + number] = (average_born_males_year[i] + average_born_females_year[i] + average_born_unknown_year[
                i]) / b_num
        if b_num == 0:
            ws['AM' + number] = 0

    # Города
    towns_list = []
    ready_towns_list = []
    count_list = []
    town = str

    for i in range(4, 700):
        number = str(i)
        if ws['B' + number].value is not None:
            town = ws['B' + number].value
            towns_list.append(town)

    for i in range(len(towns_list)):
        if towns_list[i] not in ready_towns_list:
            ready_towns_list.append(towns_list[i])

    for i in range(len(ready_towns_list)):
        count = 0
        for j in range(len(towns_list)):
            if ready_towns_list[i] == towns_list[j]:
                count += 1
        count_list.append(count)

    average_males_number = 0
    average_females_number = 0
    average_unknown_number = 0
    average_all_number = 0
    average_males_born = 0
    average_females_born = 0
    average_unknown_born = 0
    average_all_born = 0
    average_fallen = 0

    for i in range(len(years_list) - 1):
        number = i + 5
        number = str(number)
        average_males_number += ws['AF' + number].value
        average_females_number += ws['AG' + number].value
        average_unknown_number += ws['AH' + number].value
        average_all_number += ws['AI' + number].value
        average_males_born += ws['AJ' + number].value
        average_females_born += ws['AK' + number].value
        average_unknown_born += ws['AL' + number].value
        average_all_born += ws['AM' + number].value
        average_fallen += ws['AN' + number].value

    total_average_males_number = average_males_number / len(years_list)
    total_average_females_number = average_females_number / len(years_list)
    total_average_unknown_number = average_unknown_number / len(years_list)
    total_average_all_number = average_all_number / len(years_list)
    total_average_males_born = average_males_born / len(years_list)
    total_average_females_born = average_males_born / len(years_list)
    total_average_unknown_born = average_males_born / len(years_list)
    total_average_all_born = average_all_born / len(years_list)
    total_average_fallen = average_fallen / len(years_list)

    for i in range(len(ready_towns_list)):
        number = i + 5
        number = str(number)
        ws['AB' + number] = ready_towns_list[i]
        ws['AC' + number] = count_list[i]

    for i in range(len(years_list)):
        if years_list[i] == 'Всего':
            number = str(i + 5)
            ws['AF' + number] = total_average_males_number
            ws['AG' + number] = total_average_females_number
            ws['AH' + number] = total_average_unknown_number
            ws['AI' + number] = total_average_all_number
            ws['AJ' + number] = total_average_males_born
            ws['AK' + number] = total_average_females_born
            ws['AL' + number] = total_average_unknown_born
            ws['AM' + number] = total_average_all_born
            ws['AN' + number] = total_average_fallen

    # Средняя геометрическая

    # Численность
    for i in range(len(years_list) - 1):
        if i != len(years_list) - 2:
            number = str(i + 5)
            ws['AP' + number] = str(years_list[i + 1]) + '/' + str(years_list[i])

    for i in range(len(males_list) - 2):
        number = str(i + 5)
        if males_list[i] != 0:
            ws['AQ' + number] = males_list[i + 1] / males_list[i]
        else:
            ws['AQ' + number] = None

    for i in range(len(females_list) - 2):
        number = str(i + 5)
        if females_list[i] != 0:
            ws['AR' + number] = females_list[i + 1] / females_list[i]
        else:
            ws['AR' + number] = None

    for i in range(len(unknown_list) - 2):
        number = str(i + 5)
        if unknown_list[i] != 0:
            ws['AS' + number] = unknown_list[i + 1] / unknown_list[i]
        else:
            ws['AS' + number] = None

    total_number_list = []

    for i in range(len(males_list)):
        total_number = males_list[i] + females_list[i] + unknown_list[i]
        total_number_list.append(total_number)

    for i in range(len(total_number_list) - 2):
        number = str(i + 5)
        if total_number_list[i] != 0:
            ws['AT' + number] = total_number_list[i + 1] / total_number_list[i]
        else:
            ws['AT' + number] = None

    # Родившиеся
    for i in range(len(born_males_list) - 2):
        number = str(i + 5)
        if born_males_list[i] != 0:
            ws['AU' + number] = born_males_list[i + 1] / born_males_list[i]
        else:
            ws['AU' + number] = None

    for i in range(len(born_females_list) - 2):
        number = str(i + 5)
        if born_females_list[i] != 0:
            ws['AV' + number] = born_females_list[i + 1] / born_females_list[i]
        else:
            ws['AV' + number] = None

    for i in range(len(born_unknown_list) - 2):
        number = str(i + 5)
        if born_unknown_list[i] != 0:
            ws['AW' + number] = born_unknown_list[i + 1] / born_unknown_list[i]
        else:
            ws['AW' + number] = None

    total_born_list = []

    for i in range(len(born_males_list)):
        total_born = born_males_list[i] + born_females_list[i] + born_unknown_list[i]
        total_born_list.append(total_born)

    for i in range(len(total_born_list) - 2):
        number = str(i + 5)
        if total_born_list[i] != 0:
            ws['AX' + number] = total_born_list[i + 1] / total_born_list[i]
        else:
            ws['AX' + number] = None

    # Павшие
    for i in range(len(fallen_list) - 2):
        number = str(i + 5)
        if fallen_list[i] != 0:
            ws['AY' + number] = fallen_list[i + 1] / fallen_list[i]
        else:
            ws['AY' + number] = None

    # Средний темп роста

    ws['AP' + str(len(years_list) + 3)] = 'Средний темп'

    # Численность
    # Самцы
    temp = 0
    temp_list = []

    for i in range(len(years_list) - 2):
        number = str(i + 5)
        temp = ws['AQ' + number].value
        temp_list.append(temp)
    temp = 1

    for i in range(len(temp_list)):
        if temp_list[i] is not None:
            temp = temp * temp_list[i]

    n = 1 / (len(years_list) - 2)
    ws['AQ' + str(len(years_list) + 3)] = temp ** n

    # Самки
    temp = 0
    temp_list = []

    for i in range(len(years_list) - 2):
        number = str(i + 5)
        temp = ws['AR' + number].value
        temp_list.append(temp)
    temp = 1

    for i in range(len(temp_list)):
        if temp_list[i] is not None:
            temp = temp * temp_list[i]

    n = 1 / (len(years_list) - 2)
    ws['AR' + str(len(years_list) + 3)] = temp ** n

    # Всего
    temp = 0
    temp_list = []

    for i in range(len(years_list) - 2):
        number = str(i + 5)
        temp = ws['AT' + number].value
        temp_list.append(temp)
    temp = 1

    for i in range(len(temp_list)):
        if temp_list[i] is not None:
            temp = temp * temp_list[i]

    n = 1 / (len(years_list) - 2)
    ws['AT' + str(len(years_list) + 3)] = temp ** n

    # Рождаемость
    # Самцы
    temp = 0
    temp_list = []

    for i in range(len(years_list) - 2):
        number = str(i + 5)
        temp = ws['AU' + number].value
        temp_list.append(temp)
    temp = 1

    for i in range(len(temp_list)):
        if temp_list[i] is not None:
            temp = temp * temp_list[i]
    n = 1 / (len(years_list) - 2)
    ws['AU' + str(len(years_list) + 3)] = temp ** n

    # Самки
    temp = 0
    temp_list = []

    for i in range(len(years_list) - 2):
        number = str(i + 5)
        temp = ws['AV' + number].value
        temp_list.append(temp)
    temp = 1

    for i in range(len(temp_list)):
        if temp_list[i] is not None:
            temp = temp * temp_list[i]
    n = 1 / (len(years_list) - 2)
    ws['AV' + str(len(years_list) + 3)] = temp ** n

    # Всего
    temp = 0
    temp_list = []

    for i in range(len(years_list) - 2):
        number = str(i + 5)
        temp = ws['AX' + number].value
        temp_list.append(temp)
    temp = 1

    for i in range(len(temp_list)):
        if temp_list[i] is not None:
            temp = temp * temp_list[i]
    n = 1 / (len(years_list) - 2)
    ws['AX' + str(len(years_list) + 3)] = temp ** n

    # Павшие
    temp = 0
    temp_list = []

    for i in range(len(years_list) - 2):
        number = str(i + 5)
        temp = ws['AY' + number].value
        temp_list.append(temp)
    temp = 1

    for i in range(len(temp_list)):
        if temp_list[i] is not None:
            temp = temp * temp_list[i]
    n = 1 / (len(years_list) - 2)
    ws['AY' + str(len(years_list) + 3)] = temp ** n

    # Мода

    for i in range(len(years_list) - 1):
        if i != len(years_list) - 1:
            number = str(i + 5)
            ws['BA' + number] = str(years_list[i])

    m_number = 5
    for i in range(len(cell_list)):
        if cell_list[i] != 700:
            males_list = []
            females_list = []
            unknown_list = []
            all_list = []
            born_males_list = []
            born_females_list = []
            born_unknown_list = []
            born_all_list = []
            fallen_list = []
            males_mode_list = []
            females_mode_list = []
            unknown_mode_list = []
            all_mode_list = []
            born_males_mode_list = []
            born_females_mode_list = []
            born_unknown_mode_list = []
            born_all_mode_list = []
            fallen_mode_list = []

            for j in range(cell_list[i], cell_list[i + 1]):

                if ws['D' + str(j)].value is not None:
                    males_list.append(ws['D' + str(j)].value)

                if ws['E' + str(j)].value is not None:
                    females_list.append(ws['E' + str(j)].value)

                if ws['F' + str(j)].value is not None:
                    unknown_list.append(ws['F' + str(j)].value)

                if ws['G' + str(j)].value is not None:
                    all_list.append(ws['G' + str(j)].value)

                if ws['H' + str(j)].value is not None:
                    born_males_list.append(ws['H' + str(j)].value)

                if ws['I' + str(j)].value is not None:
                    born_females_list.append(ws['I' + str(j)].value)

                if ws['J' + str(j)].value is not None:
                    born_unknown_list.append(ws['J' + str(j)].value)

                if ws['K' + str(j)].value is not None:
                    born_all_list.append(ws['K' + str(j)].value)

            # Численность

            # Самцы
            for k in range(len(males_list)):
                count = 0
                for l in range(len(males_list)):
                    if males_list[k] == males_list[l]:
                        count += 1
                temp_mode = [males_list[k], count]
                males_mode_list.append(temp_mode)

            m_temp_mode = []

            for h in range(len(males_mode_list)):
                m_temp_mode.append(males_mode_list[h][1])

            if len(m_temp_mode) == 0:
                m_mode = None
            else:
                m_mode = max(m_temp_mode)

            for q in range(len(males_mode_list)):
                if males_mode_list[q][1] == m_mode:
                    ws['BB' + str(m_number)] = males_mode_list[q][0]
                    break

            # Самки
            for k in range(len(females_list)):
                count = 0
                for l in range(len(females_list)):
                    if females_list[k] == females_list[l]:
                        count += 1
                temp_mode = [females_list[k], count]
                females_mode_list.append(temp_mode)
            m_temp_mode = []

            for h in range(len(females_mode_list)):
                m_temp_mode.append(females_mode_list[h][1])

            if len(m_temp_mode) == 0:
                m_mode = None
            else:
                m_mode = max(m_temp_mode)

            for q in range(len(females_mode_list)):
                if females_mode_list[q][1] == m_mode:
                    ws['BC' + str(m_number)] = females_mode_list[q][0]
                    break

            # Неизвестно
            for k in range(len(unknown_list)):
                count = 0
                for l in range(len(unknown_list)):
                    if unknown_list[k] == unknown_list[l]:
                        count += 1
                temp_mode = [unknown_list[k], count]
                unknown_mode_list.append(temp_mode)

            m_temp_mode = []

            for h in range(len(unknown_mode_list)):
                m_temp_mode.append(unknown_mode_list[h][1])

            if len(m_temp_mode) == 0:
                m_mode = None
            else:
                m_mode = max(m_temp_mode)

            for q in range(len(unknown_mode_list)):
                if unknown_mode_list[q][1] == m_mode:
                    ws['BD' + str(m_number)] = unknown_mode_list[q][0]
                    break

            # Общее
            for k in range(len(all_list)):
                count = 0
                for l in range(len(all_list)):
                    if all_list[k] == all_list[l]:
                        count += 1
                temp_mode = [all_list[k], count]
                all_mode_list.append(temp_mode)

            m_temp_mode = []

            for h in range(len(all_mode_list)):
                m_temp_mode.append(all_mode_list[h][1])

            if len(m_temp_mode) == 0:
                m_mode = None
            else:
                m_mode = max(m_temp_mode)

            for q in range(len(all_mode_list)):
                if all_mode_list[q][1] == m_mode:
                    ws['BE' + str(m_number)] = all_mode_list[q][0]
                    break

            # Рождаемость
            # Самцы
            for k in range(len(born_males_list)):
                count = 0
                for l in range(len(born_males_list)):
                    if born_males_list[k] == born_males_list[l]:
                        count += 1
                temp_mode = []
                temp_mode.append(born_males_list[k])
                temp_mode.append(count)
                born_males_mode_list.append(temp_mode)

            m_temp_mode = []

            for h in range(len(born_males_mode_list)):
                m_temp_mode.append(born_males_mode_list[h][1])

            if len(m_temp_mode) == 0:
                m_mode = None
            else:
                m_mode = max(m_temp_mode)

            for q in range(len(born_males_mode_list)):
                if born_males_mode_list[q][1] == m_mode:
                    ws['BF' + str(m_number)] = born_males_mode_list[q][0]
                    break
            # Самки
            for k in range(len(born_females_list)):
                count = 0
                for l in range(len(born_females_list)):
                    if born_females_list[k] == born_females_list[l]:
                        count += 1
                temp_mode = []
                temp_mode.append(born_females_list[k])
                temp_mode.append(count)
                born_females_mode_list.append(temp_mode)

            m_temp_mode = []

            for h in range(len(born_females_mode_list)):
                m_temp_mode.append(born_females_mode_list[h][1])

            if len(m_temp_mode) == 0:
                m_mode = None
            else:
                m_mode = max(m_temp_mode)

            for q in range(len(born_females_mode_list)):
                if born_females_mode_list[q][1] == m_mode:
                    ws['BG' + str(m_number)] = born_females_mode_list[q][0]
                    break
            # Неизвестно
            for k in range(len(born_unknown_list)):
                count = 0
                for l in range(len(born_unknown_list)):
                    if born_unknown_list[k] == born_unknown_list[l]:
                        count += 1
                temp_mode = []
                temp_mode.append(born_unknown_list[k])
                temp_mode.append(count)
                born_unknown_mode_list.append(temp_mode)

            m_temp_mode = []

            for h in range(len(born_unknown_mode_list)):
                m_temp_mode.append(born_unknown_mode_list[h][1])

            if len(m_temp_mode) == 0:
                m_mode = None
            else:
                m_mode = max(m_temp_mode)

            for q in range(len(born_unknown_mode_list)):
                if born_unknown_mode_list[q][1] == m_mode:
                    ws['BH' + str(m_number)] = born_unknown_mode_list[q][0]
                    break
            # Общее
            for k in range(len(born_all_list)):
                count = 0
                for l in range(len(born_all_list)):
                    if born_all_list[k] == born_all_list[l]:
                        count += 1
                temp_mode = []
                temp_mode.append(born_all_list[k])
                temp_mode.append(count)
                born_all_mode_list.append(temp_mode)

            m_temp_mode = []

            for h in range(len(born_all_mode_list)):
                m_temp_mode.append(born_all_mode_list[h][1])

            if len(m_temp_mode) == 0:
                m_mode = None

            else:
                m_mode = max(m_temp_mode)

            for q in range(len(born_all_mode_list)):
                if born_all_mode_list[q][1] == m_mode:
                    ws['BI' + str(m_number)] = born_all_mode_list[q][0]
                    break
            # Павшие
            for k in range(len(fallen_list)):
                count = 0
                for l in range(len(fallen_list)):
                    if fallen_list[k] == fallen_list[l]:
                        count += 1
                temp_mode = []
                temp_mode.append(fallen_list[k])
                temp_mode.append(count)
                fallen_mode_list.append(temp_mode)

            m_temp_mode = []

            for h in range(len(fallen_mode_list)):
                m_temp_mode.append(fallen_mode_list[h][1])

            if len(m_temp_mode) == 0:
                m_mode = None
            else:
                m_mode = max(m_temp_mode)

            for q in range(len(fallen_mode_list)):
                if fallen_mode_list[q][1] == m_mode:
                    ws['BJ' + str(m_number)] = fallen_mode_list[q][0]
                    break

            m_number = int(m_number)
            m_number += 1

    # Дисперсия/Среднее квадратичное отклонение

    # По годам
    for i in range(len(years_list)):

        if i != len(years_list):
            number = str(i + 5)
            ws['BL' + number] = str(years_list[i])
            ws['BW' + number] = str(years_list[i])

    for i in range(len(cell_list)):

        if cell_list[i] != 700:
            '''males = 0
            females = 0
            unknown = 0
            all_ = 0
            born_males = 0
            born_females = 0
            born_unknown = 0
            born_all_ = 0
            fallen = 0'''
            males_list = []
            females_list = []
            unknown_list = []
            all_list = []
            born_males_list = []
            born_females_list = []
            born_unknown_list = []
            born_all_list = []
            fallen_list = []

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['D' + number].value != None:
                    males = ws['D' + number].value
                    males_list.append(males)

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['E' + number].value != None:
                    females = ws['E' + number].value
                    females_list.append(females)

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['F' + number].value != None:
                    unknown = ws['F' + number].value
                    unknown_list.append(unknown)

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['G' + number].value != None:
                    all_ = ws['G' + number].value
                    all_list.append(all_)

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['H' + number].value != None:
                    born_males = ws['H' + number].value
                    born_males_list.append(born_males)

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['I' + number].value != None:
                    born_females = ws['I' + number].value
                    born_females_list.append(born_females)

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['J' + number].value != None:
                    born_unknown = ws['J' + number].value
                    born_unknown_list.append(born_unknown)

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['K' + number].value != None:
                    born_all_ = ws['K' + number].value
                    born_all_list.append(born_all_)

            for j in range(cell_list[i], cell_list[i + 1]):
                number = str(j)
                if ws['O' + number].value != None:
                    fallen = ws['O' + number].value
                    fallen_list.append(fallen)

            number = str(i + 5)

            deviation = 0
            if len(males_list) > 1:
                for j in range(len(males_list)):
                    deviation += (males_list[j] - ws['AF' + number].value) ** 2
                ws['BM' + number] = deviation
                ws['BX' + number] = (deviation / (len(males_list) - 1)) ** (1 / 2)

            deviation = 0
            if len(females_list) > 1:
                for j in range(len(females_list)):
                    deviation += (females_list[j] - ws['AG' + number].value) ** 2
                ws['BN' + number] = deviation
                ws['BY' + number] = (deviation / (len(females_list) - 1)) ** (1 / 2)

            deviation = 0
            if len(unknown_list) > 1:
                for j in range(len(unknown_list)):
                    deviation += (unknown_list[j] - ws['AH' + number].value) ** 2
                ws['BO' + number] = deviation
                ws['BZ' + number] = (deviation / (len(unknown_list) - 1)) ** (1 / 2)

            deviation = 0
            if len(all_list) > 1:
                for j in range(len(all_list)):
                    deviation += (all_list[j] - ws['AI' + number].value) ** 2
                ws['BP' + number] = deviation
                ws['CA' + number] = (deviation / (len(all_list) - 1)) ** (1 / 2)

            deviation = 0
            if len(born_males_list) > 1:
                for j in range(len(born_males_list)):
                    deviation += (born_males_list[j] - ws['AJ' + number].value) ** 2
                ws['BQ' + number] = deviation
                ws['CB' + number] = (deviation / (len(born_males_list) - 1)) ** (1 / 2)
            deviation = 0
            if len(born_females_list) > 1:
                for j in range(len(born_females_list)):
                    deviation += (born_females_list[j] - ws['AK' + number].value) ** 2
                ws['BR' + number] = deviation
                ws['CC' + number] = (deviation / (len(born_females_list) - 1)) ** (1 / 2)

            deviation = 0
            if len(born_unknown_list) > 1:
                for j in range(len(born_unknown_list)):
                    deviation += (born_unknown_list[j] - ws['AL' + number].value) ** 2
                ws['BS' + number] = deviation
                ws['CD' + number] = (deviation / (len(born_unknown_list) - 1)) ** (1 / 2)

            deviation = 0
            if len(born_all_list) > 1:
                for j in range(len(born_all_list)):
                    deviation += (born_all_list[j] - ws['AM' + number].value) ** 2
                ws['BT' + number] = deviation
                ws['CE' + number] = (deviation / (len(born_all_list) - 1)) ** (1 / 2)

            deviation = 0
            if len(fallen_list) > 1:
                for j in range(len(fallen_list)):
                    deviation += (fallen_list[j] - ws['AN' + number].value) ** 2
                ws['BU' + number] = deviation
                ws['CF' + number] = (deviation / (len(fallen_list) - 1)) ** (1 / 2)

    # Общее
    males = 0
    females = 0
    unknown = 0
    all_ = 0
    born_males = 0
    born_females = 0
    born_unknown = 0
    born_all_ = 0
    fallen = 0
    males_list = []
    females_list = []
    unknown_list = []
    all_list = []
    born_males_list = []
    born_females_list = []
    born_unknown_list = []
    born_all_list = []
    fallen_list = []

    for j in range(3, 1000):
        number = str(j)
        if ws['D' + number].value != None:
            males = ws['D' + number].value
            males_list.append(males)

    for j in range(3, 1000):
        number = str(j)
        if ws['E' + number].value != None:
            females = ws['E' + number].value
            females_list.append(females)

    for j in range(3, 1000):
        number = str(j)
        if ws['F' + number].value != None:
            unknown = ws['F' + number].value
            unknown_list.append(unknown)

    for j in range(3, 1000):
        number = str(j)
        if ws['G' + number].value != None:
            all_ = ws['G' + number].value
            all_list.append(all_)

    for j in range(3, 1000):
        number = str(j)
        if ws['H' + number].value != None:
            born_males = ws['H' + number].value
            born_males_list.append(born_males)

    for j in range(3, 1000):
        number = str(j)
        if ws['I' + number].value != None:
            born_females = ws['I' + number].value
            born_females_list.append(born_females)
    for j in range(3, 1000):
        number = str(j)
        if ws['J' + number].value != None:
            born_unknown = ws['J' + number].value
            born_unknown_list.append(born_unknown)

    for j in range(3, 1000):
        number = str(j)
        if ws['K' + number].value != None:
            born_all_ = ws['K' + number].value
            born_all_list.append(born_all_)

    for j in range(3, 1000):
        number = str(j)
        if ws['O' + number].value != None:
            fallen = ws['O' + number].value
            fallen_list.append(fallen)

    number = str((len(years_list) + 4))

    deviation = 0
    if len(males_list) > 1:
        for i in range(len(males_list)):
            deviation += (males_list[i] - ws['AF' + number].value) ** 2
        ws['BM' + number] = deviation
        ws['BX' + number] = (deviation / (len(males_list) - 1)) ** (1 / 2)

    deviation = 0
    if len(females_list) > 1:
        for i in range(len(females_list)):
            deviation += (females_list[i] - ws['AG' + number].value) ** 2
        ws['BN' + number] = deviation
        ws['BY' + number] = (deviation / (len(females_list) - 1)) ** (1 / 2)

    deviation = 0
    if len(unknown_list) > 1:
        for i in range(len(unknown_list)):
            deviation += (unknown_list[i] - ws['AH' + number].value) ** 2
        ws['BO' + number] = deviation
        ws['BZ' + number] = (deviation / (len(unknown_list) - 1)) ** (1 / 2)

    deviation = 0
    if len(all_list) > 1:
        for i in range(len(all_list)):
            deviation += (all_list[i] - ws['AI' + number].value) ** 2
        ws['BP' + number] = deviation
        ws['CA' + number] = (deviation / (len(all_list) - 1)) ** (1 / 2)

    deviation = 0
    if len(born_males_list) > 1:
        for i in range(len(born_males_list)):
            deviation += (born_males_list[i] - ws['AJ' + number].value) ** 2
        ws['BQ' + number] = deviation
        ws['CB' + number] = (deviation / (len(born_males_list) - 1)) ** (1 / 2)

    deviation = 0
    if len(born_females_list) > 1:
        for i in range(len(born_females_list)):
            deviation += (born_females_list[i] - ws['AK' + number].value) ** 2
        ws['BR' + number] = deviation
        ws['CC' + number] = (deviation / (len(born_females_list) - 1)) ** (1 / 2)

    deviation = 0
    if len(born_unknown_list) > 1:
        for i in range(len(born_unknown_list)):
            deviation += (born_unknown_list[i] - ws['AL' + number].value) ** 2
        ws['BS' + number] = deviation
        ws['CD' + number] = (deviation / (len(born_unknown_list) - 1)) ** (1 / 2)

    deviation = 0
    if len(born_all_list) > 1:
        for i in range(len(born_all_list)):
            deviation += (born_all_list[i] - ws['AM' + number].value) ** 2
        ws['BT' + number] = deviation
        ws['CE' + number] = (deviation / (len(born_all_list) - 1)) ** (1 / 2)

    deviation = 0
    if len(fallen_list) > 1:
        for i in range(len(fallen_list)):
            deviation += (fallen_list[i] - ws['AN' + number].value) ** 2
        ws['BU' + number] = deviation
        ws['CF' + number] = (deviation / (len(fallen_list) - 1)) ** (1 / 2)

    # Коэффициент вариации

    count = 0
    for i in range(len(years_list)):
        if i != len(years_list):
            number = str(i + 5)
            ws['CH' + number] = str(years_list[i])

            if ws['BX' + number].value != None and ws['AF' + number].value != 0:
                ws['CI' + number] = (100 * ws['BX' + number].value) / ws['AF' + number].value

            if ws['BY' + number].value != None and ws['AG' + number].value != 0:
                ws['CJ' + number] = (100 * ws['BY' + number].value) / ws['AG' + number].value

            if ws['BZ' + number].value != None and ws['AH' + number].value != 0:
                ws['CK' + number] = (100 * ws['BZ' + number].value) / ws['AH' + number].value

            if ws['CA' + number].value != None and ws['AI' + number].value != 0:
                ws['CL' + number] = (100 * ws['CA' + number].value) / ws['AI' + number].value

            if ws['CB' + number].value != None and ws['AJ' + number].value != 0:
                ws['CM' + number] = (100 * ws['CB' + number].value) / ws['AJ' + number].value

            if ws['CC' + number].value != None and ws['AK' + number].value != 0:
                ws['CN' + number] = (100 * ws['CC' + number].value) / ws['AK' + number].value

            if ws['CD' + number].value != None and ws['AL' + number].value != 0:
                ws['CO' + number] = (100 * ws['CD' + number].value) / ws['AL' + number].value

            if ws['CE' + number].value != None and ws['AM' + number].value != 0:
                ws['CP' + number] = (100 * ws['CE' + number].value) / ws['AM' + number].value

            if ws['CF' + number].value != None and ws['AN' + number].value != 0:
                ws['CQ' + number] = (100 * ws['CF' + number].value) / ws['AN' + number].value

        count = number

    # Ошибка

    deviation = 0
    for i in range(1, 72):
        number = str(i + 4)
        ws['CS' + number] = deviation

        males_count = 0
        for j in range(len(males_list)):
            if deviation == 0:
                if males_list[j] == ws['AF' + count].value:
                    males_count += 1
            if deviation != 0:
                if ws['AF' + count].value < males_list[j] <= (
                        ws['AF' + count].value + (ws['BX' + count].value * deviation)):
                    males_count += 1
                if ws['AF' + count].value > males_list[j] >= (
                        ws['AF' + count].value - (ws['BX' + count].value * deviation)):
                    males_count += 1
        if len(males_list) != 0:
            ws['CT' + number] = males_count / len(males_list) * 100

        females_count = 0
        for j in range(len(females_list)):
            if deviation == 0:
                if females_list[j] == ws['AG' + count].value:
                    females_count += 1
            if deviation != 0:
                if ws['AG' + count].value < females_list[j] <= (
                        ws['AG' + count].value + (ws['BY' + count].value * deviation)):
                    females_count += 1
                if ws['AG' + count].value > females_list[j] >= (
                        ws['AG' + count].value - (ws['BY' + count].value * deviation)):
                    females_count += 1
        if len(females_list) != 0:
            ws['CU' + number] = females_count / len(females_list) * 100

        unknown_count = 0
        for j in range(len(unknown_list)):
            if deviation == 0:
                if unknown_list[j] == ws['AH' + count].value:
                    unknown_count += 1
            if deviation != 0:
                if ws['AH' + count].value < unknown_list[j] <= (
                        ws['AH' + count].value + (ws['BZ' + count].value * deviation)):
                    unknown_count += 1
                if ws['AH' + count].value > unknown_list[j] >= (
                        ws['AH' + count].value - (ws['BZ' + count].value * deviation)):
                    unknown_count += 1
        if len(unknown_list) != 0:
            ws['CV' + number] = unknown_count / len(unknown_list) * 100

        all_count = 0
        for j in range(len(all_list)):
            if deviation == 0:
                if all_list[j] == ws['AI' + count].value:
                    all_count += 1
            if deviation != 0:
                if ws['AI' + count].value < all_list[j] <= (
                        ws['AI' + count].value + (ws['CA' + count].value * deviation)):
                    all_count += 1
                if ws['AI' + count].value > all_list[j] >= (
                        ws['AI' + count].value - (ws['CA' + count].value * deviation)):
                    all_count += 1
        if len(all_list) != 0:
            ws['CW' + number] = all_count / len(all_list) * 100

        born_males_count = 0
        for j in range(len(born_males_list)):
            if deviation == 0:
                if born_males_list[j] == ws['AJ' + count].value:
                    born_males_count += 1
            if deviation != 0:
                if ws['AJ' + count].value < born_males_list[j] <= (
                        ws['AJ' + count].value + (ws['CB' + count].value * deviation)):
                    born_males_count += 1
                if ws['AJ' + count].value > born_males_list[j] >= (
                        ws['AJ' + count].value - (ws['CB' + count].value * deviation)):
                    born_males_count += 1
        if len(born_males_list) != 0:
            ws['CX' + number] = born_males_count / len(born_males_list) * 100

        born_females_count = 0
        for j in range(len(born_females_list)):
            if deviation == 0:
                if born_females_list[j] == ws['AK' + count].value:
                    born_females_count += 1
            if deviation != 0:
                if ws['AK' + count].value < born_females_list[j] <= (
                        ws['AK' + count].value + (ws['CC' + count].value * deviation)):
                    born_females_count += 1
                if ws['AK' + count].value > born_females_list[j] >= (
                        ws['AK' + count].value - (ws['CC' + count].value * deviation)):
                    born_females_count += 1
        if len(born_females_list) != 0:
            ws['CY' + number] = born_females_count / len(born_females_list) * 100

        born_unknown_count = 0
        for j in range(len(born_unknown_list)):
            if deviation == 0:
                if born_unknown_list[j] == ws['AL' + count].value:
                    born_unknown_count += 1
            if deviation != 0:
                if ws['AL' + count].value < born_unknown_list[j] <= (
                        ws['AL' + count].value + (ws['CD' + count].value * deviation)):
                    born_unknown_count += 1
                if ws['AL' + count].value > born_unknown_list[j] >= (
                        ws['AL' + count].value - (ws['CD' + count].value * deviation)):
                    born_unknown_count += 1
        if len(born_unknown_list) != 0:
            ws['CZ' + number] = born_unknown_count / len(born_unknown_list) * 100

        born_all_count = 0
        for j in range(len(born_all_list)):
            if deviation == 0:
                if born_all_list[j] == ws['AM' + count].value:
                    born_all_count += 1
            if deviation != 0:
                if ws['AM' + count].value < born_all_list[j] <= (
                        ws['AM' + count].value + (ws['CE' + count].value * deviation)):
                    born_all_count += 1
                if ws['AM' + count].value > born_all_list[j] >= (
                        ws['AM' + count].value - (ws['CE' + count].value * deviation)):
                    born_all_count += 1
        if len(born_all_list) != 0:
            ws['DA' + number] = born_all_count / len(born_all_list) * 100

        fallen_count = 0
        for j in range(len(fallen_list)):
            if deviation == 0:
                if fallen_list[j] == ws['AN' + count].value:
                    fallen_count += 1
            if deviation != 0:
                if ws['AN' + count].value < fallen_list[j] <= (
                        ws['AN' + count].value + (ws['CF' + count].value * deviation)):
                    fallen_count += 1
                if ws['AN' + count].value > fallen_list[j] >= (
                        ws['AN' + count].value - (ws['CF' + count].value * deviation)):
                    fallen_count += 1
        if len(fallen_list) != 0:
            ws['DB' + number] = fallen_count / len(fallen_list) * 100

        deviation += 0.1

    wb.save(name + '.xlsx')


root = Tk()
root.title("Импорт и статистика")

root.geometry("+700+400")

root.minsize(400, 150)
root.maxsize(400, 150)

root.attributes("-toolwindow", True)

label_1 = ttk.Label(text="Введите наименование вида")
label_1.place(x=30, y=30)
entry_1 = ttk.Entry()
entry_1.place(x=220, y=30, width=145)
label_2 = ttk.Label(text="Укажите временной диапазон")
label_2.place(x=30, y=60)
entry_2 = ttk.Entry()
entry_2.place(x=220, y=60, width=145)
btn = ttk.Button(text="Расчет", command=click)
btn.place(x=165, y=90)

root.protocol("WM_DELETE_WINDOW", finish)

root.mainloop()
